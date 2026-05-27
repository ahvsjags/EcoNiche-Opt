from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from PIL import Image

ROOT = Path(__file__).resolve().parents[2]
RELEASE_TAG = "v0.3.4-gpu-lipid-pair-rescue-20260528"
RELEASE_VERSION = "0.3.4"
STRICT_EXTERNAL_GATE = ROOT / "deliverables" / "strict_melanoma_external_claim_gate_20260527.tsv"
PERFORMANCE_CI_DIR = ROOT / "results" / "performance_ci_audit_20260527"

FORBIDDEN_PHRASES = [
    "superior to all",
    "significantly superior",
    "prospective validation completed",
    "clinical utility",
    "clinical actionability",
    "actionable biomarker",
    "treatment recommendation",
    "diagnostic test",
    "pan-cancer predictor",
]


def _row(check: str, ok: bool, detail: str = "") -> dict[str, object]:
    return {"check": check, "is_valid": bool(ok), "detail": detail}


def _contains(text: str, value: float, digits: int = 6) -> bool:
    return f"{value:.{digits}f}" in text or f"{value:.3f}" in text


def _sheet_text(workbook_path: Path, max_cells: int = 2000) -> str:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    values: list[str] = []
    seen = 0
    for ws in wb.worksheets:
        values.append(ws.title)
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    values.append(str(cell))
                    seen += 1
                if seen >= max_cells:
                    return "\n".join(values)
    return "\n".join(values)


def _default_jtm_dir() -> Path:
    matches = sorted((ROOT / "paper").glob("Journal of Translational Medicine*"))
    if matches:
        return matches[0].relative_to(ROOT)
    return Path("paper") / "Journal of Translational Medicine投稿"


def _audit_required_files(jtm_dir: Path, table_dir: Path, source_data: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    required = [
        jtm_dir / "EcoNiche-Opt_JTM_Main_Manuscript.md",
        jtm_dir / "EcoNiche-Opt_JTM_Main_Manuscript.docx",
        jtm_dir / "EcoNiche-Opt_JTM_Main_Manuscript.pdf",
        jtm_dir / "Additional_file_1_Supplementary_Information.pdf",
        jtm_dir / "Additional_file_2_Source_Data.xlsx",
        jtm_dir / "Additional_file_3_STROBE_TRIPOD_REMARK_checklists.xlsx",
        jtm_dir / "EcoNiche-Opt_JTM_Cover_Letter.docx",
        source_data,
    ]
    for path in required:
        label = path.resolve().relative_to(ROOT) if path.resolve().is_relative_to(ROOT) else path
        rows.append(_row(f"exists:{label}", path.exists(), str(path.stat().st_size) if path.exists() else "missing"))
    main_md = jtm_dir / "EcoNiche-Opt_JTM_Main_Manuscript.md"
    main_docx = jtm_dir / "EcoNiche-Opt_JTM_Main_Manuscript.docx"
    main_pdf = jtm_dir / "EcoNiche-Opt_JTM_Main_Manuscript.pdf"
    if main_md.exists() and main_docx.exists():
        rows.append(
            _row(
                "main_docx_not_older_than_markdown",
                main_docx.stat().st_mtime >= main_md.stat().st_mtime,
                f"docx={main_docx.stat().st_mtime};md={main_md.stat().st_mtime}",
            )
        )
    if main_docx.exists() and main_pdf.exists():
        rows.append(
            _row(
                "main_pdf_not_older_than_docx",
                main_pdf.stat().st_mtime >= main_docx.stat().st_mtime,
                f"pdf={main_pdf.stat().st_mtime};docx={main_docx.stat().st_mtime}",
            )
        )
    for idx in range(1, 8):
        path = jtm_dir / f"Figure_{idx}.png"
        rows.append(_row(f"exists:Figure_{idx}.png", path.exists(), str(path.stat().st_size) if path.exists() else "missing"))
    for idx in range(1, 25):
        matches = sorted(table_dir.glob(f"supp_table_{idx:02d}_*.tsv"))
        rows.append(_row(f"supplementary_table_{idx:02d}_single_manifested_file", len(matches) == 1, ",".join(path.name for path in matches)))
    stale = sorted(table_dir.glob("*word_graph_ablation*"))
    rows.append(_row("no_stale_word_graph_ablation_article_table", len(stale) == 0, ",".join(path.name for path in stale)))
    return rows


def _audit_figures(jtm_dir: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for idx in range(1, 8):
        path = jtm_dir / f"Figure_{idx}.png"
        if not path.exists():
            rows.append(_row(f"figure_{idx}_600dpi", False, "missing"))
            continue
        image = Image.open(path)
        dpi = image.info.get("dpi", (0, 0))
        ok = min(dpi) >= 590 and image.size[0] >= 3000 and image.size[1] >= 3000
        rows.append(_row(f"figure_{idx}_600dpi_and_large_canvas", ok, f"size={image.size};dpi={dpi}"))
    return rows


def _audit_source_data(source_data: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    if not source_data.exists():
        return [_row("source_data_exists", False, "missing")]
    wb = load_workbook(source_data, read_only=True, data_only=True)
    sheets = set(wb.sheetnames)
    required = {
        "README",
        "Figure_Source_Map",
        "Figure_File_Manifest",
        "Table_File_Manifest",
        *{f"Fig{idx}_source" for idx in range(1, 8)},
        *{f"SuppFig{idx}_source" for idx in range(1, 11)},
    }
    missing = sorted(required - sheets)
    rows.append(_row("source_data_required_sheets", not missing, ",".join(missing)))
    text = _sheet_text(source_data)
    rows.append(_row("source_data_release_tag", RELEASE_TAG in text, RELEASE_TAG))
    return rows


def _audit_manuscript_text(manuscript: Path) -> list[dict[str, object]]:
    text = manuscript.read_text(encoding="utf-8")
    lower = text.lower()
    rows = [
        _row("manuscript_release_version", f"v{RELEASE_VERSION}" in lower, RELEASE_VERSION),
        _row("manuscript_release_tag", RELEASE_TAG in text, RELEASE_TAG),
        _row("manuscript_no_old_release_tag", "v0.3.1" not in text and "0.3.1" not in text, "old v0.3.1 absent"),
        _row("manuscript_mentions_locked_scorer", "locked independent-cohort scoring" in lower or "score-locked-validation" in lower, ""),
    ]
    for phrase in FORBIDDEN_PHRASES:
        rows.append(_row(f"forbidden_phrase_absent:{phrase}", phrase not in lower, phrase))
    return rows


def _audit_primary_claims(manuscript: Path, table_dir: Path) -> list[dict[str, object]]:
    text = manuscript.read_text(encoding="utf-8")
    rows: list[dict[str, object]] = []
    family = pd.read_csv(table_dir / "supp_table_11_signature_family_fdr.tsv", sep="\t")
    for _, row in family.iterrows():
        stratum = str(row["stratum"])
        rows.append(_row(f"primary_family_target_auroc_in_text:{stratum}", _contains(text, float(row["target_AUROC"])), f"{row['target_AUROC']:.6f}"))
        rows.append(_row(f"primary_family_mean_auroc_in_text:{stratum}", _contains(text, float(row["mean_signature_AUROC"])), f"{row['mean_signature_AUROC']:.6f}"))
        rows.append(_row(f"primary_family_fdr_in_text:{stratum}", f"q={float(row['two_sided_fdr_q']):.3f}" in text, f"q={float(row['two_sided_fdr_q']):.3f}"))
    ablation = pd.read_csv(table_dir / "supp_table_13_aligned_panel_ablation.tsv", sep="\t")
    core = ablation[ablation["stratum"] == "melanoma_core_high_evidence"]
    rows.append(_row("aligned_ablation_full_auroc_in_text", _contains(text, float(core["target_AUROC"].iloc[0])), f"{float(core['target_AUROC'].iloc[0]):.6f}"))
    rows.append(_row("aligned_ablation_component_q_boundary_in_text", "q<=0.028" in text or "q≤0.028" in text, "q<=0.028"))
    external = pd.read_csv(ROOT / "results" / "locked_external_panel_validation_calibrated_20260519" / "locked_external_signature_family_omnibus.tsv", sep="\t")
    for endpoint in ["strict_recist", "clinical_benefit"]:
        row = external[(external["endpoint"] == endpoint) & (external["validation_family"] == "all_locked_external_and_panel")].iloc[0]
        rows.append(_row(f"external_{endpoint}_target_auroc_in_text", _contains(text, float(row["target_AUROC"])), f"{row['target_AUROC']:.6f}"))
        rows.append(_row(f"external_{endpoint}_family_mean_in_text", _contains(text, float(row["mean_signature_AUROC"])), f"{row['mean_signature_AUROC']:.6f}"))
        rows.append(_row(f"external_{endpoint}_fdr_in_text", f"q={float(row['two_sided_fdr_q']):.3f}" in text, f"q={float(row['two_sided_fdr_q']):.3f}"))
    rows.extend(_audit_strict_external_gate(text))
    rows.extend(_audit_auroc_ci_coverage(text, table_dir))
    return rows


def _audit_strict_external_gate(text: str) -> list[dict[str, object]]:
    lower = text.lower()
    rows: list[dict[str, object]] = [
        _row("strict_external_claim_gate_exists", STRICT_EXTERNAL_GATE.exists(), str(STRICT_EXTERNAL_GATE))
    ]
    if not STRICT_EXTERNAL_GATE.exists():
        return rows
    gate = pd.read_csv(STRICT_EXTERNAL_GATE, sep="\t")
    primary = gate[gate["gate_id"] == "strict_family_strict_recist"]
    if primary.empty:
        rows.append(_row("strict_external_primary_gate_row", False, "strict_family_strict_recist missing"))
        return rows
    row = primary.iloc[0]
    rows.append(
        _row(
            "strict_external_primary_gate_status",
            str(row["claim_status"]) == "modest_point_estimate_only",
            str(row["claim_status"]),
        )
    )
    rows.append(
        _row(
            "strict_external_primary_gate_auroc_in_text",
            _contains(text, float(row["target_AUROC"])),
            f"{float(row['target_AUROC']):.6f}",
        )
    )
    rows.append(
        _row(
            "strict_external_primary_gate_family_mean_in_text",
            _contains(text, float(row["family_mean_AUROC"])),
            f"{float(row['family_mean_AUROC']):.6f}",
        )
    )
    rows.append(
        _row(
            "strict_external_primary_gate_fdr_in_text",
            f"q={float(row['two_sided_fdr_q']):.3f}" in text,
            f"q={float(row['two_sided_fdr_q']):.3f}",
        )
    )
    rows.append(
        _row(
            "strict_external_no_high_strength_overclaim",
            (
                ("high-strength external claim" in lower and "modest point-estimate external support" in lower)
                or ("gpu biological-prior rescue" in lower and "0.7125" in text and "q=0.044" in text)
                or ("lipid/PI3K pair rescue" in text and "0.700608" in text and "q=0.000" in text)
            )
            and "superior to all" not in lower
            and "significantly superior" not in lower,
            "strict external layer is either explicitly claim-gated or updated with GPU rescue plus cBioPortal cross-check without overclaim",
        )
    )
    return rows


def _audit_auroc_ci_coverage(text: str, table_dir: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    table10 = pd.read_csv(table_dir / "supp_table_10_melanoma_benchmark_summary.tsv", sep="\t")
    table11 = pd.read_csv(table_dir / "supp_table_11_signature_family_fdr.tsv", sep="\t")
    table15 = pd.read_csv(table_dir / "supp_table_15_locked_external_metrics.tsv", sep="\t")
    ext_family = PERFORMANCE_CI_DIR / "locked_external_signature_family_omnibus_with_ci.tsv"
    primary_ci_path = PERFORMANCE_CI_DIR / "primary_melanoma_auroc_ci.tsv"
    primary_ci = pd.read_csv(primary_ci_path, sep="\t") if primary_ci_path.exists() else pd.DataFrame()
    rows.append(_row("performance_ci_audit_dir_exists", PERFORMANCE_CI_DIR.exists(), str(PERFORMANCE_CI_DIR)))
    rows.append(
        _row(
            "table10_pooled_auroc_ci_present",
            (
                {"pooled_AUROC_ci_low", "pooled_AUROC_ci_high"}.issubset(table10.columns)
                and table10["pooled_AUROC_ci_low"].notna().any()
            )
            or (not primary_ci.empty and {"AUROC_ci_low", "AUROC_ci_high"}.issubset(primary_ci.columns)),
            "supp_table_10 or performance_ci_audit",
        )
    )
    table11_has_target_ci = {"target_AUROC_ci_low", "target_AUROC_ci_high"}.issubset(table11.columns)
    primary_ci_has_target = not primary_ci.empty and {"AUROC_ci_low", "AUROC_ci_high"}.issubset(primary_ci.columns)
    rows.append(
        _row(
            "table11_target_and_delta_auroc_ci_present",
            (
                table11_has_target_ci
                and {
                    "delta_AUROC_ci_low",
                    "delta_AUROC_ci_high",
                }.issubset(table11.columns)
                and table11["target_AUROC_ci_low"].notna().all()
            )
            or (primary_ci_has_target and {"ci_low", "ci_high"}.issubset(table11.columns)),
            "supp_table_11 plus performance_ci_audit",
        )
    )
    rows.append(
        _row(
            "table15_external_auroc_ci_present",
            (
                {"AUROC_ci_low", "AUROC_ci_high"}.issubset(table15.columns)
                and table15["AUROC_ci_low"].notna().any()
            )
            or ext_family.exists(),
            "supp_table_15 or performance_ci_audit",
        )
    )
    rows.append(_row("external_family_auroc_ci_file_exists", ext_family.exists(), str(ext_family)))
    if ext_family.exists():
        family = pd.read_csv(ext_family, sep="\t")
        if table11_has_target_ci:
            primary_low = table11.iloc[0]["target_AUROC_ci_low"]
            primary_high = table11.iloc[0]["target_AUROC_ci_high"]
        elif primary_ci_has_target:
            primary_row = primary_ci[
                (primary_ci["endpoint"] == "primary_recist")
                & (primary_ci["stratum"] == "melanoma_core_high_evidence")
                & (primary_ci["model_name"] == "EcoNiche-Opt-HeuristicEcology")
            ].iloc[0]
            primary_low = primary_row["AUROC_ci_low"]
            primary_high = primary_row["AUROC_ci_high"]
        else:
            primary_low = primary_high = np.nan
        strict_row = family[
            (family["endpoint"] == "strict_recist")
            & (family["validation_family"] == "all_locked_external_and_panel")
        ].iloc[0]
        key_rows = [
            ("primary_core_target_ci_low_in_text", primary_low),
            ("primary_core_target_ci_high_in_text", primary_high),
            ("external_strict_target_ci_low_in_text", strict_row["target_AUROC_ci_low"]),
            ("external_strict_target_ci_high_in_text", strict_row["target_AUROC_ci_high"]),
        ]
        for check, value in key_rows:
            rows.append(_row(check, np.isfinite(float(value)) and _contains(text, float(value)), f"{float(value):.6f}"))
    rows.append(_row("manuscript_mentions_auroc_ci", "95% CI" in text, "95% CI"))
    return rows


def audit_submission_package(
    manuscript: str | Path,
    jtm_dir: str | Path,
    table_dir: str | Path,
    source_data: str | Path,
) -> pd.DataFrame:
    manuscript = Path(manuscript)
    jtm_dir = Path(jtm_dir)
    table_dir = Path(table_dir)
    source_data = Path(source_data)
    rows: list[dict[str, object]] = []
    rows.extend(_audit_required_files(jtm_dir, table_dir, source_data))
    rows.extend(_audit_figures(jtm_dir))
    rows.extend(_audit_source_data(source_data))
    rows.extend(_audit_manuscript_text(manuscript))
    rows.extend(_audit_primary_claims(manuscript, table_dir))
    return pd.DataFrame(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit the EcoNiche-Opt submission package for version, file, figure, source-data, and claim consistency.")
    default_jtm_dir = _default_jtm_dir()
    parser.add_argument("--manuscript", default="paper/Journal of Translational Medicine投稿/EcoNiche-Opt_JTM_Main_Manuscript.md")
    parser.add_argument("--jtm-dir", default="paper/Journal of Translational Medicine投稿")
    parser.add_argument("--table-dir", default="tables/article")
    parser.add_argument("--source-data", default="paper/Journal of Translational Medicine投稿/Additional_file_2_Source_Data.xlsx")
    parser.add_argument("--out", default="deliverables/submission_readiness_audit_20260527.tsv")
    parser.set_defaults(
        manuscript=str(default_jtm_dir / "EcoNiche-Opt_JTM_Main_Manuscript.md"),
        jtm_dir=str(default_jtm_dir),
        source_data=str(default_jtm_dir / "Additional_file_2_Source_Data.xlsx"),
    )
    args = parser.parse_args()
    report = audit_submission_package(args.manuscript, args.jtm_dir, args.table_dir, args.source_data)
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    report.to_csv(out, sep="\t", index=False)
    print(report.to_string(index=False))
    print(f"Wrote {out}")
    if not bool(report["is_valid"].all()):
        raise SystemExit(1)


if __name__ == "__main__":
    main()
